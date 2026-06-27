"""Extract (service_name, price, duration) rows from uploaded price-list files.

Supports the formats clinics actually publish openly (TZ §3.1): Excel (.xlsx),
CSV, PDF, DOCX. The goal is heuristic robustness, not perfection: a clinic price
list is almost always a 2-column "name / price" table, sometimes with a duration
or category column. We detect the price column (cells that parse as money) and the
name column (the widest text column), so the same logic works across layouts.

Returns a list of (name: str, price: float, duration_days: int | None).
"""
from __future__ import annotations

import csv
import logging
import re
from pathlib import Path

logger = logging.getLogger(__name__)

ExtractedRow = tuple[str, float, "int | None"]

_PRICE_HINT = re.compile(r"цен|сто[ие]м|price|тенге|kzt|₸|тг", re.IGNORECASE)
_NAME_HINT = re.compile(r"наимен|услуг|анализ|иссл|service|name|назван", re.IGNORECASE)
_DUR_HINT = re.compile(r"срок|дн|готов|duration|term", re.IGNORECASE)
_DURATION_RE = re.compile(r"(\d+)")
# A money cell: at least 3 digits (clinic prices are ≥100 KZT), spaces/dots/commas allowed.
_MONEY_RE = re.compile(r"^\s*\d[\d\s.,  ]{2,}\s*(?:₸|тг|тенге|kzt|руб)?\s*$", re.IGNORECASE)


def parse_price(text: str) -> float | None:
    """'3 980 ₸' / '8 400' / '1 234,50' / '1.234,00' -> float, else None."""
    if text is None:
        return None
    s = str(text).strip()
    if not s:
        return None
    # keep digits and separators only
    s = re.sub(r"[^\d.,]", "", s.replace(" ", "").replace(" ", "").replace(" ", ""))
    if not s:
        return None
    # If both separators present, assume the last one is the decimal sep.
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        # comma as decimal only if exactly 2 trailing digits, else thousands sep
        s = s.replace(",", ".") if re.search(r",\d{2}$", s) else s.replace(",", "")
    digits = re.sub(r"[^\d.]", "", s)
    if not digits or digits == ".":
        return None
    try:
        val = float(digits)
    except ValueError:
        return None
    # Reject implausible prices (a stray year, a phone fragment, etc.)
    return val if 50 <= val <= 100_000_000 else None


def parse_duration(text: str) -> int | None:
    if not text:
        return None
    m = _DURATION_RE.search(str(text))
    return int(m.group(1)) if m else None


def _looks_like_money(cell: str) -> bool:
    return bool(cell) and (_MONEY_RE.match(str(cell)) is not None) and parse_price(cell) is not None


def _rows_from_matrix(matrix: list[list[str]]) -> list[ExtractedRow]:
    """Turn a 2D grid (from a sheet / CSV / extracted table) into price rows.

    Strategy: find the price column by header hint, else by 'most cells parse as
    money'. Name column = header hint, else the widest text column to its left.
    """
    matrix = [[("" if c is None else str(c)).strip() for c in row] for row in matrix if any(row)]
    if not matrix:
        return []

    ncols = max(len(r) for r in matrix)
    grid = [r + [""] * (ncols - len(r)) for r in matrix]

    # --- header detection (optional) ---
    price_col = name_col = dur_col = None
    header_rows = 0
    for hr in grid[:2]:
        joined = " | ".join(hr).lower()
        if _PRICE_HINT.search(joined) or _NAME_HINT.search(joined):
            for ci, cell in enumerate(hr):
                if price_col is None and _PRICE_HINT.search(cell):
                    price_col = ci
                if name_col is None and _NAME_HINT.search(cell):
                    name_col = ci
                if dur_col is None and _DUR_HINT.search(cell):
                    dur_col = ci
            header_rows += 1
            break

    body = grid[header_rows:] if header_rows else grid

    # --- price column by money density (fallback) ---
    if price_col is None:
        best_ci, best_hits = None, 0
        for ci in range(ncols):
            hits = sum(1 for r in body if _looks_like_money(r[ci]))
            if hits > best_hits:
                best_ci, best_hits = ci, hits
        if best_ci is None or best_hits < max(2, len(body) // 5):
            return []
        price_col = best_ci

    # --- name column = widest text column that isn't the price column ---
    if name_col is None or name_col == price_col:
        best_ci, best_len = None, 0
        for ci in range(ncols):
            if ci == price_col:
                continue
            total = sum(len(r[ci]) for r in body if not _looks_like_money(r[ci]))
            if total > best_len:
                best_ci, best_len = ci, total
        name_col = best_ci if best_ci is not None else (0 if price_col != 0 else 1 if ncols > 1 else 0)

    out: list[ExtractedRow] = []
    for r in body:
        name = r[name_col].strip() if name_col < len(r) else ""
        price = parse_price(r[price_col]) if price_col < len(r) else None
        if not name or price is None or len(name) < 3:
            continue
        dur = parse_duration(r[dur_col]) if (dur_col is not None and dur_col < len(r)) else None
        out.append((re.sub(r"\s+", " ", name), price, dur))
    return out


# --- per-format readers ------------------------------------------------------
def _from_xlsx(path: str) -> list[ExtractedRow]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    rows: list[ExtractedRow] = []
    for ws in wb.worksheets:
        matrix = [[c for c in row] for row in ws.iter_rows(values_only=True)]
        rows.extend(_rows_from_matrix(matrix))
    wb.close()
    return rows


def _from_csv(path: str) -> list[ExtractedRow]:
    for enc in ("utf-8-sig", "cp1251", "utf-8"):
        try:
            with open(path, newline="", encoding=enc) as f:
                sample = f.read(4096)
                f.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
                except csv.Error:
                    dialect = csv.excel
                matrix = [list(r) for r in csv.reader(f, dialect)]
            return _rows_from_matrix(matrix)
        except (UnicodeDecodeError, csv.Error):
            continue
    return []


def _from_pdf(path: str) -> list[ExtractedRow]:
    import pdfplumber

    rows: list[ExtractedRow] = []
    _LINE = re.compile(r"^(.+?)[\s.…]{1,}([\d\s  .,]{3,})\s*(?:₸|тг|тенге)?\s*$")
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables() or []:
                rows.extend(_rows_from_matrix(table))
            # fallback: "Service name .... 3 980 ₸" lines when there are no tables
            if not (page.extract_tables() or []):
                for line in (page.extract_text() or "").splitlines():
                    m = _LINE.match(line.strip())
                    if not m:
                        continue
                    name, price = m.group(1).strip(" .…"), parse_price(m.group(2))
                    if name and price is not None and len(name) >= 3:
                        rows.append((re.sub(r"\s+", " ", name), price, None))
    return rows


def _from_docx(path: str) -> list[ExtractedRow]:
    from docx import Document

    doc = Document(path)
    rows: list[ExtractedRow] = []
    for table in doc.tables:
        matrix = [[cell.text for cell in row.cells] for row in table.rows]
        rows.extend(_rows_from_matrix(matrix))
    return rows


_READERS = {
    "xlsx": _from_xlsx, "xlsm": _from_xlsx, "xls": _from_xlsx,
    "csv": _from_csv, "pdf": _from_pdf, "docx": _from_docx,
}
SUPPORTED_EXTENSIONS = sorted(_READERS)


def extract_price_rows(path: str) -> list[ExtractedRow]:
    ext = Path(path).suffix.lower().lstrip(".")
    reader = _READERS.get(ext)
    if reader is None:
        raise ValueError(f"Unsupported file type: .{ext} (supported: {', '.join(SUPPORTED_EXTENSIONS)})")
    rows = reader(path)
    # de-dup identical (name, price) pairs that some layouts repeat
    seen: set[tuple[str, float]] = set()
    uniq: list[ExtractedRow] = []
    for name, price, dur in rows:
        key = (name.lower(), price)
        if key in seen:
            continue
        seen.add(key)
        uniq.append((name, price, dur))
    logger.info("[file_extract] %s -> %d price rows", path, len(uniq))
    return uniq
